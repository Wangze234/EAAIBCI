# -*- coding: utf-8 -*-
"""
Created on Wed Sep 26 20:35:24 2018


@author: HAN_RUIZHI yb77447@umac.mo OR  501248792@qq.com

This code is the first version of BLS Python.
If you have any questions about the code or find any bugs
   or errors during use, please feel free to contact me.
If you have any questions about the original paper,
   please contact the authors of related paper.
"""
import openpyxl

'''

Installation has been tested with Python 3.5.
Since the package is written in python 3.5, 
python 3.5 with the pip tool must be installed first. 
It uses the following dependencies: numpy(1.16.3), scipy(1.2.1), keras(2.2.0), sklearn(0.20.3)  
You can install these packages first, by the following commands:

pip install numpy
pip install scipy
pip install keras (if use keras data_load())
pip install scikit-learn
'''
import torch
import csv
from snntorch import spikegen
from sklearn import preprocessing
import numpy as np
import scipy.io as scio
from BroadLearningSystem import BLS, BLS_AddEnhanceNodes, BLS_AddFeatureEnhanceNodes
from sklearn.preprocessing import OneHotEncoder
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt



def read_label_values(file_path, column_index):
    # file_path 表示label文件路径
    # column_index 表示所在列
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    # 选择第一个工作表
    sheet = wb.active
    # 创建一个空数组，用于存储列值
    column_values = []
    # 遍历指定列的每一行，将值添加到数组中
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_index,
                               max_col=column_index, values_only=True):
        column_values.append(row[0])
    # 关闭 Excel 文件
    wb.close()
    return column_values

def read_data_values(file_path):
    datas = []
    with open(file_path, "r") as csvfile:
        csvreader = csv.reader(csvfile)
        for row in csvreader:
            datas.append(row)
    # 关闭文档
    csvfile.close()
    datas = np.double(datas)
    # BCIdata = BCIdata.astype(float)
    print(datas.shape)
    return datas

def  sort_label(tag, label):
    if tag == 'Train':
        labels_indices = np.argmax(label, axis=1)
        sorted_indices = np.argsort(labels_indices)
        labels = label[sorted_indices]
    else:
        sorted_indices = np.argsort(label)
        labels = label[sorted_indices]
        # 将labels转化为独热编码 一位是1，其他位都是0的编码
        enc = OneHotEncoder()
        labels = enc.fit_transform(labels.reshape(-1, 1)).toarray()
    return labels

print("***************************************读取训练集*******************************************")
tag = 'Train'
# 读取训练集标签
TrainLabelDataFile = 'F:\\EAAIBCI\\EAAIBCI\\Training set\\Data1.mat'
# 读取排序后提取的训练集特征
TrainDataFile = 'F:\\EAAIBCI\\EAAIBCI\\TrainFeatures\\Normalizedfeatures1.csv'
TrainLabelData = scio.loadmat(TrainLabelDataFile)
TrainLabelData = TrainLabelData['epo_train']
TrainLabel = np.double(TrainLabelData[0][0][5])
TrainLabel = np.transpose(TrainLabel)
# 使用相同的索引对标签进行排序
Train_labels = sort_label(tag, TrainLabel)
# 获取数据
Train_datas = read_data_values(TrainDataFile)


print("***************************************读取测试集*******************************************")
tag = 'Test'
# 读取测试集标签
TestLabelDataFile = 'F:\\EAAIBCI\\EAAIBCI\\Test set\\Testlabel.xlsx'
# 读取排序后提取的测试集特征
TestDataFile = 'F:\\EAAIBCI\\EAAIBCI\\TestFeatures\\Normalizedfeatures01.csv'
# 获取原始数据中的label
TestLabel = read_label_values(TestLabelDataFile, 0)
TestLabel = np.array(TestLabel)
# 使用相同的索引对标签进行排序
Test_labels = sort_label(tag, TestLabel)
Test_datas = read_data_values(TestDataFile)
# 先划分训练集和测试集
# traindata, testdata, trainlabel, testlabel = train_test_split(Train_datas, Train_labels, test_size=0.1,)


N1 =10 #  # of nodes belong to each window
N2 =10 #  # of windows -------Feature mapping layer
N3 =500  # of enhancement nodes -----Enhance layer
L = 5    #  # of incremental steps
M1 = 30  #  # of adding enhance nodes
s = 0.8  #  shrink coefficient
C = 2**-30 # Regularization coefficient

print('-------------------BLS_BASE---------------------------')
B1,B1test,B1test1,test_acc1=BLS(Train_datas, Train_labels, Test_datas, Test_labels, s, C, N1, N2, N3)









