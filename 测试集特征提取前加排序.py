import numpy as np
# 获取排序后的索引
import scipy.stats as sp
import scipy.io as scio
import csv
import pandas as pd
import openpyxl
from scipy.signal import welch


def read_label_values(file_path, column_index):
    # file_path 表示label文件路径
    # column_index 表示所在列
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    # 选择第一个工作表
    sheet = wb.active
    # 创建一个空数组，用于存储列值
    column_values = []
    # 遍历指定列的每一行，将值添加到数组中
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_index,
                               max_col=column_index, values_only=True):
        column_values.append(row[0])
    # 关闭 Excel 文件
    wb.close()
    return column_values


def read_EEG_values(file_path):
    # 获取预处理滤波后的数据
    processdata = scio.loadmat(file_path)
    processdata = processdata['EEG']
    traindataProc = np.double(processdata[0][0][15])
    # 获取滤波后的脑电数据
    arr_transposed = np.transpose(traindataProc, (1, 0, 2))
    return arr_transposed


def sortdata_bylabel(label, data):  # 按标签大小对数据排序
    sorted_indices = np.argsort(label)
    sorted_data = data[:, :, sorted_indices]
    return sorted_data


def hjorth(input):  # function for hjorth
    realinput = input
    hjorth_activity = np.zeros((3, 6))
    hjorth_mobility = np.zeros((3, 6))
    hjorth_diffmobility = np.zeros((3, 6))
    hjorth_complexity = np.zeros((3, 6))
    for i in range(3):
        for j in range(6):
            hjorth_activity[i, j] = np.var(realinput[i * 265:(i + 1) * 265, j])
    diff_input = np.diff(realinput, axis=0)
    diff_diffinput = np.diff(diff_input, axis=0)
    for i in range(3):
        for j in range(6):
            hjorth_mobility[i, j] = np.sqrt(
                np.var(diff_input[i * 265:(i + 1) * 265, j]) / hjorth_activity[i, j])
            hjorth_diffmobility[i, j] = np.sqrt(
                np.var(diff_diffinput[i * 264:(i + 1) * 264, j]) / np.var(
                    diff_input[i * 265:(i + 1) * 265, j]))
            hjorth_complexity[i, j] = hjorth_diffmobility[i, j] / hjorth_mobility[i, j]
    return np.sum(hjorth_activity, axis=0) / 8, np.sum(hjorth_mobility, axis=0) / 8, np.sum(
        hjorth_complexity, axis=0) / 8

# 提取对应维度的特征
def Get_Features(dim, data):

    features2 = []
    for i in range(dim):
        hjorthfeatures = hjorth(data[:, :, i])
        combined_features = np.hstack((hjorthfeatures))
        features2.append(combined_features)

    features2 = np.array(features2)
    print(features2.shape)
    lines = [l for l in features2]

    for i in range(len(lines[0]) - 1):
        columns = []
        for j in range(0, len(lines)):
            columns.append(float(lines[j][i]))
        mean = np.mean(columns, axis=0)
        std_dev = np.std(columns, axis=0)

        for j in range(0, len(lines)):
            lines[j][i] = (float(lines[j][i]) - mean) / std_dev
    lines = np.array(lines)
    print(lines.shape)
    return lines

def main():
    # 包含标签的原始训练集数据
    rawdataFile = 'F:\\EAAIBCI\\EAAIBCI\\Test set\\Testlabel.xlsx'
    # eeglab处理好之后的训练集数据
    procdatafile = 'F:\\EAAIBCI\\EAAIBCI\\TestdataAfilter\\Test set_filter\\Su15.mat'
    # 获取原始数据中的label
    label_data = read_label_values(rawdataFile, 14)
    # 获取预处理滤波后的数据
    arr_transposed = read_EEG_values(procdatafile)
    # 根据整数标签对数据和标签数组进行排序
    sorted_data = sortdata_bylabel(label_data, arr_transposed)
    # 由第三维度对一二维矩阵提取特征
    third_dim_size = sorted_data.shape[2]
    # 获取特征
    features = Get_Features(third_dim_size, sorted_data)
    # 保存特征
    writer = csv.writer(
        open('TestFeatures/Normalizedfeatures15.csv', 'w',
             newline=''))  # This file will store the normalized features
    writer.writerows(features)


main()
